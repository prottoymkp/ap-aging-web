import datetime as dt
import unittest
from io import BytesIO

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook

from engine import find_header_row, fifo_aging, parse_top_sheet, transform_ap_ledger


class TestFifoAging(unittest.TestCase):
    def test_find_header_row_ignores_float_cells_while_matching_labels(self):
        rows = [
            [np.nan, 42.0, None],
            [1.0, "SL.No", "Date", "Debit Amount", "Credit Amount", 5.5],
        ]

        header_idx, col_map = find_header_row(rows)

        self.assertEqual(header_idx, 1)
        self.assertEqual(col_map["date"], 2)
        self.assertEqual(col_map["debit"], 3)
        self.assertEqual(col_map["credit"], 4)

    def test_parse_top_sheet_ignores_float_cells_while_matching_header(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Top Sheet"
        ws.append([np.nan, 100.5, None])
        ws.append(["Supplier Name", "Materials Value", "Paid Amount", "Unpaid Amount/Liabilities"])
        ws.append(["ABC Traders", 1000, 250, 750])
        buffer = BytesIO()
        wb.save(buffer)

        parsed = parse_top_sheet(buffer.getvalue(), "Top Sheet")

        self.assertEqual(parsed.loc[0, "Supplier Name"], "ABC Traders")
        self.assertEqual(parsed.loc[0, "Materials Value"], 1000)
        self.assertEqual(parsed.loc[0, "Paid Amount"], 250)
        self.assertEqual(parsed.loc[0, "Unpaid Amount/Liabilities"], 750)

    def test_transform_handles_empty_optional_output_sheets(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Top Sheet"
        ws.append(["Supplier Name", "Materials Value", "Paid Amount", "Unpaid Amount/Liabilities"])
        ws.append(["ABC Traders", 1000, 250, 750])

        supplier = wb.create_sheet("ABC")
        supplier["A1"] = "ABC Traders"
        supplier.append([])
        supplier.append(["SL.No", "Date", "Remarks", "Debit Amount", "Credit Amount"])
        supplier.append([1, dt.date(2024, 1, 5), "Invoice", 0, 1000])
        supplier.append([2, dt.date(2024, 1, 15), "Payment", 250, 0])
        supplier.append(["Total Amount"])

        buffer = BytesIO()
        wb.save(buffer)

        out_bytes = transform_ap_ledger(buffer.getvalue(), dt.date(2024, 1, 31), "Top Sheet")
        out_wb = load_workbook(BytesIO(out_bytes), read_only=True)

        self.assertIn("Aging_Summary", out_wb.sheetnames)
        self.assertIn("Undated_Entries", out_wb.sheetnames)

    def test_balanced_with_future_dated_row_results_in_zero_outstanding(self):
        as_of = dt.date(2024, 1, 31)
        tx_df = pd.DataFrame(
            [
                {"date": dt.date(2024, 1, 10), "credit": 100.0, "debit": 0.0, "net": 100.0},
                {"date": dt.date(2024, 1, 15), "credit": 0.0, "debit": 100.0, "net": -100.0},
                # Future-dated row where credit and debit offset (net=0), so no unpaid liability remains.
                {"date": dt.date(2024, 2, 10), "credit": 50.0, "debit": 50.0, "net": 0.0},
            ]
        )

        total_payable, total_paid, balance, _, _, bs, _ = fifo_aging(tx_df, as_of)

        self.assertEqual(total_payable, 150.0)
        self.assertEqual(total_paid, 150.0)
        self.assertEqual(balance, 0.0)
        self.assertEqual(bs["future_dated_unpaid"], 0.0)
        self.assertEqual(bs["advance_overpaid"], 0.0)

    def test_true_overpayment_sets_advance_equal_to_absolute_negative_balance(self):
        as_of = dt.date(2024, 1, 31)
        tx_df = pd.DataFrame(
            [
                {"date": dt.date(2024, 1, 10), "credit": 50.0, "debit": 0.0, "net": 50.0},
                {"date": dt.date(2024, 1, 20), "credit": 0.0, "debit": 80.0, "net": -80.0},
            ]
        )

        _, _, balance, _, _, bs, _ = fifo_aging(tx_df, as_of)

        self.assertLess(balance, 0.0)
        self.assertEqual(bs["advance_overpaid"], abs(balance))
        self.assertEqual(bs["0-30"], 0.0)
        self.assertEqual(bs["31-60"], 0.0)
        self.assertEqual(bs["61-90"], 0.0)
        self.assertEqual(bs["91-180"], 0.0)
        self.assertEqual(bs["181-365"], 0.0)
        self.assertEqual(bs[">365"], 0.0)
        self.assertEqual(bs["future_dated_unpaid"], 0.0)
        self.assertEqual(bs["unknown_date_unpaid"], 0.0)

    def test_future_unpaid_liability_populates_future_bucket_without_advance(self):
        as_of = dt.date(2024, 1, 31)
        tx_df = pd.DataFrame(
            [
                {"date": dt.date(2024, 1, 10), "credit": 20.0, "debit": 0.0, "net": 20.0},
                {"date": dt.date(2024, 2, 5), "credit": 120.0, "debit": 0.0, "net": 120.0},
                {"date": dt.date(2024, 1, 12), "credit": 0.0, "debit": 20.0, "net": -20.0},
            ]
        )

        _, _, balance, _, _, bs, _ = fifo_aging(tx_df, as_of)

        self.assertGreater(balance, 0.0)
        self.assertEqual(bs["future_dated_unpaid"], 120.0)
        self.assertEqual(bs["advance_overpaid"], 0.0)

        bucket_total = (
            bs["0-30"]
            + bs["31-60"]
            + bs["61-90"]
            + bs["91-180"]
            + bs["181-365"]
            + bs[">365"]
            + bs["future_dated_unpaid"]
            + bs["unknown_date_unpaid"]
            - bs["advance_overpaid"]
        )
        self.assertEqual(balance, bucket_total)


if __name__ == "__main__":
    unittest.main()
