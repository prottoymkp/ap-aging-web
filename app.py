import datetime as dt
from io import BytesIO

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from engine import transform_ap_ledger

st.set_page_config(page_title="AP Ledger Transformer", layout="centered")
st.title("AP Ledger Transformer")
st.caption("Upload AP ledger → Run → Download aging report + see red flags on screen")

uploaded = st.file_uploader("Upload AP ledger Excel (.xlsx)", type=["xlsx"])
if uploaded is None:
    st.stop()

excel_bytes = uploaded.getvalue()

# Detect sheets so user can pick top summary tab (people rename it)
try:
    wb = load_workbook(BytesIO(excel_bytes), data_only=True)
    sheetnames = wb.sheetnames
except Exception:
    st.error("This file cannot be opened as an Excel workbook. Upload a valid .xlsx.")
    st.stop()

default_idx = sheetnames.index("Top Sheet") if "Top Sheet" in sheetnames else 0
top_sheet_name = st.selectbox("Top summary sheet tab", sheetnames, index=default_idx)

as_of = st.date_input("As-of date (aging cutoff)", value=dt.date.today())
show_diag = st.checkbox("Show diagnostics after run", value=True)

run = st.button("Run aging analysis", type="primary")
if not run:
    st.stop()

with st.spinner("Processing..."):
    try:
        out_bytes = transform_ap_ledger(
            excel_bytes=excel_bytes,
            as_of=as_of,
            top_sheet_name=top_sheet_name,
        )
    except Exception as e:
        st.error("Processing failed.")
        st.code(str(e))
        st.stop()

st.success("Done.")

st.download_button(
    "Download aging report (.xlsx)",
    data=out_bytes,
    file_name=f"AP_Aging_Report_{as_of.isoformat()}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

if not show_diag:
    st.stop()

# ---- Diagnostics (read generated output) ----
bio = BytesIO(out_bytes)
aging = pd.read_excel(bio, sheet_name="Aging_Summary", engine="openpyxl")
bio.seek(0)
mapping = pd.read_excel(bio, sheet_name="Supplier_Mapping", engine="openpyxl")
bio.seek(0)
dq = pd.read_excel(bio, sheet_name="Data_Quality_Issues", engine="openpyxl") if "Data_Quality_Issues" in load_workbook(bio, read_only=True).sheetnames else pd.DataFrame()

# Portfolio totals
st.subheader("Portfolio totals")
money_cols = [
    "total_payable", "total_paid", "balance",
    "0-30", "31-60", "61-90", "91-180", "181-365", ">365",
    "future_dated_unpaid", "unknown_date_unpaid", "advance_overpaid",
]
present_money_cols = [c for c in money_cols if c in aging.columns]
totals = aging[present_money_cols].sum(numeric_only=True).to_frame("Total").reset_index().rename(columns={"index": "metric"})
st.dataframe(totals, use_container_width=True)

# Red flags
st.subheader("Red flags (fix these in the ledger)")

flags = []

# 1) Recon delta (aging buckets don't reconcile to balance)
if "recon_delta" in aging.columns:
    flags.append(aging.loc[aging["recon_delta"].abs() > 1, ["supplier", "sheet", "recon_delta"]].assign(flag="recon_delta != 0"))

# 2) Missing dates / unknown-date unpaid
if "missing_date_rows" in aging.columns:
    flags.append(aging.loc[aging["missing_date_rows"] > 0, ["supplier", "sheet", "missing_date_rows"]].assign(flag="missing dates"))
if "unknown_date_unpaid" in aging.columns:
    flags.append(aging.loc[aging["unknown_date_unpaid"] > 0, ["supplier", "sheet", "unknown_date_unpaid"]].assign(flag="unknown-date unpaid"))

# 3) Future-dated
if "future_dated_unpaid" in aging.columns:
    flags.append(aging.loc[aging["future_dated_unpaid"] > 0, ["supplier", "sheet", "future_dated_unpaid"]].assign(flag="future-dated invoices"))

# 4) Top sheet mismatch (balance diff)
if "diff_balance" in aging.columns:
    flags.append(aging.loc[aging["diff_balance"].abs() > 1, ["supplier", "sheet", "diff_balance"]].assign(flag="Top Sheet mismatch (balance)"))

# 5) Low-confidence supplier mapping
low_map = mapping.loc[mapping["match_score"] < 0.75, ["sheet", "sheet_title", "mapped_supplier", "match_score", "method"]].copy()
low_map["flag"] = "low-confidence supplier mapping"

flag_df = pd.concat(flags, ignore_index=True) if flags else pd.DataFrame()
if not flag_df.empty:
    st.dataframe(flag_df.sort_values(["flag", "supplier", "sheet"]), use_container_width=True)
else:
    st.write("No red flags detected by the rules above.")

if not low_map.empty:
    st.subheader("Low-confidence supplier mapping (review names)")
    st.dataframe(low_map.sort_values("match_score"), use_container_width=True)

# Data quality sheet preview
if dq is not None and not dq.empty:
    st.subheader("Data Quality Issues (from report)")
    st.dataframe(dq, use_container_width=True)
