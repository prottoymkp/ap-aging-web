from __future__ import annotations

import datetime as dt
import re
from difflib import SequenceMatcher
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter


# Maintain overrides here as you discover naming mismatches.
ALIAS_OVERRIDES: Dict[str, str] = {
    # "Sheet Tab Name": "Top Sheet Supplier Name"
    # "RSF": "RSF Steel Craft",
}

STOPWORDS = {
    "ltd", "limited", "enterprise", "production", "traders", "trading",
    "corporation", "corp", "company", "co", "workshop", "printing",
    "packaging", "accessories", "and", "the"
}


# ----------------------------
# Parsing helpers
# ----------------------------

def norm_name(x: Any) -> str:
    if x is None:
        return ""
    s = str(x).lower()
    s = re.sub(r"[\.\,\-\(\)\&\/]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def normalize_tokens(x: Any) -> List[str]:
    toks = [t for t in norm_name(x).split() if t and t not in STOPWORDS]
    return toks


def match_score(a: str, b: str) -> float:
    na, nb = norm_name(a), norm_name(b)
    r = SequenceMatcher(None, na, nb).ratio()

    ta, tb = set(normalize_tokens(a)), set(normalize_tokens(b))
    j = (len(ta & tb) / len(ta | tb)) if (ta | tb) else 0.0

    bonus = 0.0
    ta_list, tb_list = normalize_tokens(a), normalize_tokens(b)
    if ta_list and tb_list and ta_list[0] == tb_list[0]:
        bonus = 0.15

    return 0.55 * j + 0.45 * r + bonus


def best_match_top(candidate: str, top_names: List[str]) -> Tuple[str, float]:
    best_name = candidate
    best = -1.0
    for t in top_names:
        sc = match_score(candidate, t)
        if sc > best:
            best = sc
            best_name = t
    return best_name, best


def parse_amount(x: Any) -> float:
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return 0.0
    if isinstance(x, (int, float, np.integer, np.floating)):
        return float(x)

    s = str(x).strip()
    if not s:
        return 0.0
    s = s.replace(",", "")
    s = re.sub(r"[^0-9\.\-\+]", "", s)
    try:
        return float(s) if s else 0.0
    except Exception:
        return 0.0


def parse_date_value(val: Any) -> Tuple[Optional[dt.date], Optional[str]]:
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return None, None

    if isinstance(val, dt.datetime):
        return val.date(), str(val)
    if isinstance(val, dt.date):
        return val, str(val)

    if isinstance(val, (int, float, np.integer, np.floating)):
        f = float(val)
        if 20000 < f < 60000:  # typical Excel serial date range
            base = dt.date(1899, 12, 30)
            try:
                return base + dt.timedelta(days=f), str(val)
            except Exception:
                return None, str(val)
        return None, str(val)

    s = str(val).strip()
    if not s:
        return None, s

    ts = pd.to_datetime(s, dayfirst=True, errors="coerce")
    if pd.isna(ts):
        return None, s
    return ts.date(), s


def is_valid_date(d: Any) -> bool:
    return isinstance(d, dt.date) and not pd.isna(d)


# ----------------------------
# Top sheet
# ----------------------------

def parse_top_sheet(excel_bytes: bytes, top_sheet_name: str) -> pd.DataFrame:
    raw = pd.read_excel(BytesIO(excel_bytes), sheet_name=top_sheet_name, header=None, engine="openpyxl")
    hdr = None
    for i in range(len(raw)):
        row = raw.iloc[i].astype(str).str.lower().tolist()
        if any("supplier name" in x for x in row):
            hdr = i
            break
    if hdr is None:
        raise ValueError(f"Could not find 'Supplier Name' header row in '{top_sheet_name}'.")

    df = pd.read_excel(BytesIO(excel_bytes), sheet_name=top_sheet_name, header=hdr, engine="openpyxl")

    def find_col(predicate) -> Optional[str]:
        for c in df.columns:
            if predicate(str(c).strip().lower()):
                return c
        return None

    c_sup = find_col(lambda s: "supplier" in s and "name" in s)
    c_mat = find_col(lambda s: "materials" in s and "value" in s)
    c_paid = find_col(lambda s: "paid" in s and "amount" in s)
    c_unpaid = find_col(lambda s: ("unpaid" in s) or ("liabilit" in s))

    if not all([c_sup, c_mat, c_paid, c_unpaid]):
        raise ValueError(f"Top sheet columns not found in '{top_sheet_name}'. Found: {list(df.columns)}")

    out = df[[c_sup, c_mat, c_paid, c_unpaid]].copy()
    out.columns = ["Supplier Name", "Materials Value", "Paid Amount", "Unpaid Amount/Liabilities"]
    out = out[out["Supplier Name"].notna()].reset_index(drop=True)

    out["Supplier Name"] = out["Supplier Name"].astype(str).str.strip()
    for c in ["Materials Value", "Paid Amount", "Unpaid Amount/Liabilities"]:
        out[c] = pd.to_numeric(out[c], errors="coerce").fillna(0.0)

    return out


# ----------------------------
# Supplier sheet extraction
# ----------------------------

def sheet_rows_values(ws) -> List[List[Any]]:
    return [
        list(r)
        for r in ws.iter_rows(
            min_row=1, max_row=ws.max_row,
            min_col=1, max_col=ws.max_column,
            values_only=True
        )
    ]


def find_header_row(rows: List[List[Any]], max_scan: int = 80) -> Tuple[Optional[int], Dict[str, int]]:
    for i, row in enumerate(rows[:max_scan]):
        texts = [str(c).strip().lower() if c is not None else "" for c in row]
        joined = " | ".join(texts)
        if (("sl.no" in joined) or ("sl no" in joined)) and ("date" in joined) and ("debit" in joined) and ("credit" in joined):
            col_map: Dict[str, int] = {}
            for j, t in enumerate(texts):
                if t.startswith("date"):
                    col_map["date"] = j
                elif "debit" in t:
                    col_map["debit"] = j
                elif "credit" in t:
                    col_map["credit"] = j
                elif "remark" in t:
                    col_map["remarks"] = j
            return i, col_map
    return None, {}


def is_termination_row(row: List[Any]) -> bool:
    for c in row:
        if isinstance(c, str):
            s = c.strip().lower()
            if (
                s.startswith("total amount")
                or s.startswith("total")
                or "total unpaid" in s
                or s.startswith("c. total unpaid")
                or "grand total" in s
            ):
                return True
    return False


def extract_transactions(ws, as_of: dt.date) -> Tuple[str, pd.DataFrame, Dict[str, Any], List[Dict[str, Any]]]:
    rows = sheet_rows_values(ws)

    title_cell = rows[0][0] if rows and rows[0] else None
    supplier_title = title_cell.strip() if isinstance(title_cell, str) and title_cell.strip() else ws.title

    header_idx, col_map = find_header_row(rows)
    if header_idx is None or "debit" not in col_map or "credit" not in col_map:
        issues = {
            "no_header": True,
            "missing_date_rows": 0,
            "negative_credit_rows": 0,
            "negative_debit_rows": 0,
            "both_debit_credit_rows": 0,
            "future_dated_any_rows": 0,
            "future_dated_any_examples": [],
            "older_dated_after_newer_year_rows": 0,
            "older_year_examples": [],
        }
        return supplier_title, pd.DataFrame(), issues, []

    date_col = col_map.get("date", 1)
    debit_col = col_map["debit"]
    credit_col = col_map["credit"]
    remarks_col = col_map.get("remarks", None)

    tx_rows: List[Dict[str, Any]] = []
    undated_rows: List[Dict[str, Any]] = []

    issues: Dict[str, Any] = {
        "no_header": False,
        "missing_date_rows": 0,
        "negative_credit_rows": 0,
        "negative_debit_rows": 0,
        "both_debit_credit_rows": 0,
        "future_dated_any_rows": 0,
        "future_dated_any_examples": [],
        "older_dated_after_newer_year_rows": 0,
        "older_year_examples": [],
    }

    max_year_seen: Optional[int] = None
    started = False
    blank_streak = 0

    for ridx in range(header_idx + 1, len(rows)):
        row = rows[ridx]
        if is_termination_row(row):
            break

        date_val = row[date_col] if date_col < len(row) else None
        debit = parse_amount(row[debit_col] if debit_col < len(row) else None)
        credit = parse_amount(row[credit_col] if credit_col < len(row) else None)

        if (date_val is None or (isinstance(date_val, float) and np.isnan(date_val))) and debit == 0 and credit == 0:
            blank_streak += 1
            if started and blank_streak >= 5:
                break
            continue
        blank_streak = 0
        started = True

        d, date_raw = parse_date_value(date_val)
        remarks = row[remarks_col] if (remarks_col is not None and remarks_col < len(row)) else None

        if debit != 0 and credit != 0:
            issues["both_debit_credit_rows"] += 1
        if debit < 0:
            issues["negative_debit_rows"] += 1
        if credit < 0:
            issues["negative_credit_rows"] += 1

        if d is None and (debit != 0 or credit != 0):
            issues["missing_date_rows"] += 1
            undated_rows.append({
                "row_index": ridx + 1,
                "date_raw": date_raw,
                "debit": debit,
                "credit": credit,
                "remarks": remarks,
            })

        if d is not None:
            if d > as_of and (debit != 0 or credit != 0):
                issues["future_dated_any_rows"] += 1
                if len(issues["future_dated_any_examples"]) < 3:
                    issues["future_dated_any_examples"].append((ridx + 1, d.isoformat(), debit, credit))

            if max_year_seen is None:
                max_year_seen = d.year
            else:
                if d.year < max_year_seen:
                    issues["older_dated_after_newer_year_rows"] += 1
                    if len(issues["older_year_examples"]) < 3:
                        issues["older_year_examples"].append((ridx + 1, d.isoformat(), max_year_seen))
                max_year_seen = max(max_year_seen, d.year)

        net = credit - debit
        tx_rows.append({
            "row_index": ridx + 1,
            "date_raw": date_raw,
            "date": pd.Timestamp(d) if d is not None else pd.NaT,
            "debit": debit,
            "credit": credit,
            "remarks": remarks,
            "net": net,
        })

    tx_df = pd.DataFrame(tx_rows)
    return supplier_title, tx_df, issues, undated_rows


# ----------------------------
# Aging logic
# ----------------------------

def bucket_from_age(age_days: int) -> str:
    if age_days <= 30:
        return "0-30"
    if age_days <= 60:
        return "31-60"
    if age_days <= 90:
        return "61-90"
    if age_days <= 180:
        return "91-180"
    if age_days <= 365:
        return "181-365"
    return ">365"


def fifo_aging(tx_df: pd.DataFrame, as_of: dt.date):
    buckets = ["0-30", "31-60", "61-90", "91-180", "181-365", ">365", "future_dated_unpaid", "unknown_date_unpaid", "advance_overpaid"]
    bs = {k: 0.0 for k in buckets}

    if tx_df is None or tx_df.empty:
        return 0.0, 0.0, 0.0, pd.NaT, pd.NaT, bs, []

    total_payable = float(tx_df["credit"].sum())
    total_paid = float(tx_df["debit"].sum())
    balance = total_payable - total_paid

    tx = tx_df.copy()
    tx["d"] = pd.to_datetime(tx["date"], errors="coerce").dt.date

    pay_mask = (tx["debit"] > 0) | (tx["net"] < 0)
    pay_dates = [d for d in tx.loc[pay_mask, "d"].tolist() if is_valid_date(d) and d <= as_of]
    last_payment = pd.Timestamp(max(pay_dates)) if pay_dates else pd.NaT

    inv_dated: List[Dict[str, Any]] = []
    inv_future: List[Dict[str, Any]] = []
    inv_undated: List[Dict[str, Any]] = []
    reductions: List[float] = []

    for _, r in tx.iterrows():
        net = float(r["net"])
        d = r["d"]
        if net > 1e-9:
            if not is_valid_date(d):
                inv_undated.append({"date": None, "amt": net})
            else:
                if d > as_of:
                    inv_future.append({"date": d, "amt": net})
                else:
                    inv_dated.append({"date": d, "amt": net})
        elif net < -1e-9:
            reductions.append(-net)

    inv_dated.sort(key=lambda x: x["date"])

    def apply_to_queue(queue: List[Dict[str, Any]], amt: float) -> float:
        idx = 0
        while amt > 1e-9 and idx < len(queue):
            take = min(queue[idx]["amt"], amt)
            queue[idx]["amt"] -= take
            amt -= take
            if queue[idx]["amt"] <= 1e-9:
                idx += 1
            else:
                break
        while queue and queue[0]["amt"] <= 1e-9:
            queue.pop(0)
        return amt

    advance = 0.0
    for red in reductions:
        rem = apply_to_queue(inv_dated, red)

        i = 0
        while rem > 1e-9 and i < len(inv_undated):
            take = min(inv_undated[i]["amt"], rem)
            inv_undated[i]["amt"] -= take
            rem -= take
            if inv_undated[i]["amt"] <= 1e-9:
                i += 1
            else:
                break
        inv_undated = [x for x in inv_undated if x["amt"] > 1e-9]

        if rem > 1e-9:
            advance += rem

    outstanding_detail: List[Dict[str, Any]] = []

    dated_out = [x for x in inv_dated if x["amt"] > 1e-9]
    future_out = [x for x in inv_future if x["amt"] > 1e-9]
    undated_out = [x for x in inv_undated if x["amt"] > 1e-9]

    dated_dates = [x["date"] for x in (dated_out + future_out) if is_valid_date(x["date"])]
    oldest_unpaid = pd.Timestamp(min(dated_dates)) if dated_dates else pd.NaT

    for it in dated_out:
        age = (as_of - it["date"]).days
        b = bucket_from_age(age)
        amt = float(it["amt"])
        bs[b] += amt
        outstanding_detail.append({
            "invoice_date": pd.Timestamp(it["date"]),
            "amount_outstanding": amt,
            "age_days": age,
            "bucket": b,
        })

    for it in future_out:
        age = (as_of - it["date"]).days
        amt = float(it["amt"])
        bs["future_dated_unpaid"] += amt
        outstanding_detail.append({
            "invoice_date": pd.Timestamp(it["date"]),
            "amount_outstanding": amt,
            "age_days": age,
            "bucket": "future",
        })

    for it in undated_out:
        amt = float(it["amt"])
        bs["unknown_date_unpaid"] += amt
        outstanding_detail.append({
            "invoice_date": pd.NaT,
            "amount_outstanding": amt,
            "age_days": np.nan,
            "bucket": "unknown_date_unpaid",
        })

    if advance > 1e-9:
        bs["advance_overpaid"] = float(advance)  # positive in summary
        outstanding_detail.append({
            "invoice_date": pd.NaT,
            "amount_outstanding": -float(advance),  # negative in detail
            "age_days": np.nan,
            "bucket": "advance_overpaid",
        })

    return total_payable, total_paid, balance, last_payment, oldest_unpaid, bs, outstanding_detail


# ----------------------------
# Output formatting
# ----------------------------

def autosize_columns(ws, max_width: int = 55) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            v = cell.value
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def format_numbers(ws, numeric_cols: List[str], date_cols: List[str]) -> None:
    header = [c.value for c in ws[1]]
    col_index = {str(h).strip(): i + 1 for i, h in enumerate(header) if h is not None}

    for name in numeric_cols:
        if name in col_index:
            c = col_index[name]
            for r in range(2, ws.max_row + 1):
                ws.cell(r, c).number_format = "#,##0"

    for name in date_cols:
        if name in col_index:
            c = col_index[name]
            for r in range(2, ws.max_row + 1):
                ws.cell(r, c).number_format = "yyyy-mm-dd"


def apply_borders(ws) -> None:
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value is None:
                continue
            cell.border = border


# ----------------------------
# Public API: transform bytes -> bytes
# ----------------------------

def transform_ap_ledger(excel_bytes: bytes, as_of: dt.date, top_sheet_name: str = "Top Sheet") -> bytes:
    # Load workbook for sheet traversal + selection
    wb = load_workbook(BytesIO(excel_bytes), data_only=True)

    if top_sheet_name not in wb.sheetnames:
        # fallback: first sheet
        top_sheet_name = wb.sheetnames[0]

    top_df = parse_top_sheet(excel_bytes, top_sheet_name)
    top_names = top_df["Supplier Name"].tolist()

    aging_rows: List[Dict[str, Any]] = []
    outstanding_rows: List[Dict[str, Any]] = []
    undated_rows: List[Dict[str, Any]] = []
    dq_rows: List[Dict[str, Any]] = []
    mapping_rows: List[Dict[str, Any]] = []

    for sname in wb.sheetnames:
        if sname == top_sheet_name:
            continue

        ws = wb[sname]
        sheet_title, tx_df, issues, und = extract_transactions(ws, as_of)

        # Supplier mapping (override -> fuzzy -> fallback)
        if sname in ALIAS_OVERRIDES:
            supplier = ALIAS_OVERRIDES[sname]
            score = 1.0
            method = "override(sheet)"
        elif sheet_title in ALIAS_OVERRIDES:
            supplier = ALIAS_OVERRIDES[sheet_title]
            score = 1.0
            method = "override(title)"
        else:
            b1, sc1 = best_match_top(sheet_title, top_names)
            b2, sc2 = best_match_top(sname, top_names)
            supplier, score, method = (b1, sc1, "fuzzy(title)") if sc1 >= sc2 else (b2, sc2, "fuzzy(sheet)")
            if score < 0.60:
                supplier = sheet_title
                method += " (low_confidence_fallback)"

        mapping_rows.append({
            "sheet": sname,
            "sheet_title": sheet_title,
            "mapped_supplier": supplier,
            "match_score": score,
            "method": method,
        })

        total_payable, total_paid, balance, last_payment, oldest_unpaid, bs, out = fifo_aging(tx_df, as_of)

        bucket_total = (
            bs["0-30"] + bs["31-60"] + bs["61-90"] + bs["91-180"] + bs["181-365"] + bs[">365"]
            + bs["future_dated_unpaid"] + bs["unknown_date_unpaid"]
            - bs["advance_overpaid"]
        )
        recon_delta = balance - bucket_total

        aging_rows.append({
            "supplier": supplier,
            "sheet": sname,
            "total_payable": total_payable,
            "total_paid": total_paid,
            "balance": balance,
            "last_payment": last_payment,
            "oldest_unpaid": oldest_unpaid,
            "0-30": bs["0-30"],
            "31-60": bs["31-60"],
            "61-90": bs["61-90"],
            "91-180": bs["91-180"],
            "181-365": bs["181-365"],
            ">365": bs[">365"],
            "future_dated_unpaid": bs["future_dated_unpaid"],
            "unknown_date_unpaid": bs["unknown_date_unpaid"],
            "advance_overpaid": bs["advance_overpaid"],
            "missing_date_rows": issues.get("missing_date_rows", 0),
            "neg_entries": (
                issues.get("negative_credit_rows", 0) + issues.get("negative_debit_rows", 0)
            ) if not issues.get("no_header") else np.nan,
            "recon_delta": recon_delta,
        })

        for u in und:
            u.update({"supplier": supplier, "sheet": sname})
            undated_rows.append(u)

        for o in out:
            o.update({"supplier": supplier, "sheet": sname})
            outstanding_rows.append(o)

        parts: List[str] = []
        if issues.get("no_header"):
            parts.append("no header detected")
        if issues.get("missing_date_rows"):
            parts.append(f"missing date rows: {issues['missing_date_rows']}")
        if issues.get("older_dated_after_newer_year_rows"):
            ex = "; ".join([f"row {r} {d} after year {y}" for (r, d, y) in issues.get("older_year_examples", [])])
            parts.append(
                f"older-dated entry after newer year: {issues['older_dated_after_newer_year_rows']} row(s) e.g. {ex}"
                if ex else
                f"older-dated entry after newer year: {issues['older_dated_after_newer_year_rows']} row(s)"
            )
        if issues.get("both_debit_credit_rows"):
            parts.append(f"both debit & credit in same row: {issues['both_debit_credit_rows']}")
        if issues.get("negative_credit_rows") or issues.get("negative_debit_rows"):
            parts.append(f"negative amounts (credit:{issues.get('negative_credit_rows',0)} debit:{issues.get('negative_debit_rows',0)})")
        if issues.get("future_dated_any_rows"):
            ex = "; ".join([f"row {r} {d} (debit {deb}, credit {cred})" for (r, d, deb, cred) in issues.get("future_dated_any_examples", [])])
            parts.append(
                f"future-dated entries: {issues['future_dated_any_rows']} row(s) e.g. {ex}"
                if ex else
                f"future-dated entries: {issues['future_dated_any_rows']} row(s)"
            )
        if parts:
            dq_rows.append({"supplier": supplier, "sheet": sname, "issues": "; ".join(parts)})

    aging_df = pd.DataFrame(aging_rows).sort_values(["supplier", "sheet"], ignore_index=True)
    outstanding_df = pd.DataFrame(outstanding_rows).sort_values(["supplier", "sheet", "bucket", "invoice_date"], ignore_index=True)
    undated_df = pd.DataFrame(undated_rows).sort_values(["supplier", "sheet", "row_index"], ignore_index=True)
    dq_df = pd.DataFrame(dq_rows).sort_values(["supplier", "sheet"], ignore_index=True)
    mapping_df = pd.DataFrame(mapping_rows).sort_values(["match_score", "sheet"], ignore_index=True)

    # Write to memory (bytes)
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        aging_df.to_excel(writer, sheet_name="Aging_Summary", index=False)
        outstanding_df.to_excel(writer, sheet_name="Outstanding_Detail", index=False)
        undated_df.to_excel(writer, sheet_name="Undated_Entries", index=False)
        dq_df.to_excel(writer, sheet_name="Data_Quality_Issues", index=False)
        top_df.to_excel(writer, sheet_name="Top_Sheet", index=False)
        mapping_df.to_excel(writer, sheet_name="Supplier_Mapping", index=False)

    buffer.seek(0)
    out_wb = load_workbook(buffer)

    fmt_plan = {
        "Aging_Summary": {
            "numeric": [
                "total_payable", "total_paid", "balance",
                "0-30", "31-60", "61-90", "91-180", "181-365", ">365",
                "future_dated_unpaid", "unknown_date_unpaid", "advance_overpaid",
                "missing_date_rows", "neg_entries", "recon_delta"
            ],
            "date": ["last_payment", "oldest_unpaid"],
        },
        "Outstanding_Detail": {
            "numeric": ["amount_outstanding", "age_days"],
            "date": ["invoice_date"],
        },
        "Undated_Entries": {
            "numeric": ["row_index", "debit", "credit"],
            "date": [],
        },
        "Top_Sheet": {
            "numeric": ["Materials Value", "Paid Amount", "Unpaid Amount/Liabilities"],
            "date": [],
        },
        "Supplier_Mapping": {
            "numeric": ["match_score"],
            "date": [],
        },
    }

    for wsname, plan in fmt_plan.items():
        if wsname not in out_wb.sheetnames:
            continue
        ws = out_wb[wsname]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        format_numbers(ws, numeric_cols=plan["numeric"], date_cols=plan["date"])
        autosize_columns(ws)
        apply_borders(ws)

    out_buffer = BytesIO()
    out_wb.save(out_buffer)
    return out_buffer.getvalue()
